from rest_framework import viewsets
from .models import Maestro
from .serializers import MaestroSerializer
from rest_framework.views import APIView
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class MaestroViewSet(viewsets.ModelViewSet):
    queryset = Maestro.objects.all()
    serializer_class = MaestroSerializer
    
class ExportarMaestrosView(APIView):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maestros Registrados"

        headers = ['Username', 'Apellido Paterno', 'Apellido Materno', 'Nombre', 'Contraseña Temporal']
        ws.append(headers)

        maestros = Maestro.objects.all()

        for maestro in maestros:
            ws.append([
                maestro.id.username,
                maestro.apellido_paterno,
                maestro.apellido_materno,
                maestro.nombre,
                maestro.contrasenaTemporal or ''
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column  
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=alumnos.xlsx'

        wb.save(response)
        return response