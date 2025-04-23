from django.shortcuts import render
from rest_framework.renderers import JSONRenderer
from rest_framework import viewsets
from .models import Alumno
from .serializers import AlumnoSerializer
from clase.models import Clase
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.files.storage import default_storage
import os
from users.models import User
from django.contrib.auth.hashers import make_password
import secrets
import string
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated


def ver_asistencia(request):
    return render(request, 'inicio.html')

class AlumnoViewSet(viewsets.ModelViewSet):
    queryset = Alumno.objects.all()
    
    serializer_class = AlumnoSerializer
    
    renderer_classes = [JSONRenderer]
    
    http_method_names = ['get', 'post', 'put', 'delete']

    def get_queryset(self):
        queryset = super().get_queryset()
        clase_id = self.request.query_params.get('clase', None)
        
        if clase_id is not None:
            try:
                clase = Clase.objects.get(id=clase_id)
                queryset = queryset.filter(grupo=clase.grupo_id)
            except Clase.DoesNotExist:
                queryset = Alumno.objects.none()
        
        return queryset


def generar_contrasena_temporal(longitud=10):
    caracteres = string.ascii_letters + string.digits
    return ''.join(secrets.choice(caracteres) for _ in range(longitud))
class CargaMasivaAlumnosView(APIView):
    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response({"error": "No se recibió ningún archivo."}, status=status.HTTP_400_BAD_REQUEST)

        print("Nombre del archivo:", file.name)
        print("Contenido del archivo:")
        for chunk in file.chunks():
            print(chunk.decode(errors="ignore"))

        path = default_storage.save(f'tmp/{file.name}', file)
        tmp_file = os.path.join(default_storage.location, path)

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(tmp_file, encoding='utf-8-sig')
            else:
                df = pd.read_excel(tmp_file)

            print("DataFrame cargado:")
            print(df.head())

            df.columns = df.columns.str.strip().str.lower()
            df.dropna(subset=["nombre", "apellido_paterno", "apellido_materno"], inplace=True) 
            contrasena_temporal = generar_contrasena_temporal()
           
            for _, row in df.iterrows():
                try:
                    contrasena_temporal = generar_contrasena_temporal()
                    username = f"{row['nombre']}.{row['apellido_paterno']}".lower()
                    password = make_password(contrasena_temporal)
                    user = User.objects.create(
                        username=username,
                        password=password,
                        rol_id=3  
                    )

                    Alumno.objects.create(
                        id=user,  
                        nombre=row["nombre"],
                        apellido_paterno=row["apellido_paterno"],
                        apellido_materno=row["apellido_materno"],
                        contrasenaTemporal=contrasena_temporal,
                        grupo=None,
                        grado=None,
                        carrera_id=None,
                    )
                except Exception as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    continue

            return Response({"message": "Alumnos cargados exitosamente."}, status=status.HTTP_200_OK)
        except Exception as e:
            print("Error durante el procesamiento:", str(e)) 
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if os.path.exists(tmp_file):
                os.remove(tmp_file)
                
class ExportarAlumnosView(APIView):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alumnos Registrados"

        headers = ['Username', 'Apellido Paterno', 'Apellido Materno', 'Nombre', 'Contraseña Temporal']
        ws.append(headers)

        alumnos = Alumno.objects.all()

        for alumno in alumnos:
            ws.append([
                alumno.id.username,
                alumno.apellido_paterno,
                alumno.apellido_materno,
                alumno.nombre,
                alumno.contrasenaTemporal or ''
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column  
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=alumnos.xlsx'

        wb.save(response)
        return response